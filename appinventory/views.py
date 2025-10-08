# Django core
from django.views.generic import TemplateView
from django.db import models
from django.db.models import F, Sum, OuterRef, Subquery, Count, Max, Q
from django.db.models.deletion import ProtectedError
from django.db import IntegrityError
from django.http import HttpResponse
# Django REST Framework (DRF)
from rest_framework.exceptions import ValidationError
from rest_framework import status, viewsets
from rest_framework.generics import ListAPIView
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.authentication import TokenAuthentication
from rest_framework.decorators import permission_classes, action, api_view
from rest_framework.permissions import (
    IsAuthenticated,
    DjangoModelPermissions,
    AllowAny,
    IsAuthenticatedOrReadOnly
    )
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from utils.datatable import handle_datatable_query
# App Models
from appinventory.models import (
    Product, Stock, Warehouse, ProductCategory,
    ProductBrand, UnitOfMeasure, UnitCategory, PriceType, InventoryMovement
    )
from apptransactions.models import Document, DocumentLine, DocumentType
# Serializers
from appinventory.serializers import (
    WarehouseSerializer, ProductCategorySerializer, ProductBrandSerializer,
    ProductSerializer, UnitOfMeasureSerializer, UnitCategorySerializer,
    PriceTypeSerializer, ProductListSerializer,ProductDetailSerializer
    )


class DashboardView(TemplateView):
    template_name = "inventory/dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Productos con stock bajo el reorder_level
        low_stock_products = (
            Product.objects.filter(is_active=True)
            .annotate(
                total_stock=Subquery(
                    Stock.objects.filter(product=OuterRef('pk'))
                    .values('product')
                    .annotate(qty=Sum('quantity'))
                    .values('qty')[:1]
                )
            )
            .filter(total_stock__lt=F('reorder_level'))
        )

        # Top 5 productos más bajos en stock
        lowest_stock_products = (
            Product.objects.filter(is_active=True)
            .annotate(
                total_stock=Subquery(
                    Stock.objects.filter(product=OuterRef('pk'))
                    .values('product')
                    .annotate(qty=Sum('quantity'))
                    .values('qty')[:1]
                )
            )
            .order_by('total_stock')[:5]
        )

        # Últimos movimientos (últimos 10 documentos con líneas)
        recent_documents = (
            Document.objects.filter(is_active=True)
            .select_related('document_type', 'party')
            .prefetch_related('lines')
            .order_by('-date')[:10]
        )

        # Totales generales
        total_warehouses = Warehouse.objects.filter(is_active=True).count()
        total_products = Product.objects.filter(is_active=True).count()
        total_stock_units = Stock.objects.aggregate(total=Sum('quantity'))['total'] or 0

        context.update({
            'low_stock_products': low_stock_products,
            'lowest_stock_products': lowest_stock_products,
            'recent_documents': recent_documents,
            'total_warehouses': total_warehouses,
            'total_products': total_products,
            'total_stock_units': total_stock_units,
            'page_title': "Inventory Dashboard"
        })

        return context

class WarehouseViewSet(viewsets.ModelViewSet):
    queryset = Warehouse.objects.all()
    serializer_class = WarehouseSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['is_active']

    @action(detail=False, methods=['patch'], url_path='clear-default')
    def clear_default(self, request):
        """
        Clear all warehouses as default (set is_default=False for all)
        This is used when setting a new warehouse as default
        """
        try:
            Warehouse.objects.filter(is_default=True).update(is_default=False)
            return Response({
                'message': 'All warehouses cleared as default',
                'success': True
            })
        except Exception as e:
            return Response({
                'error': f'Error clearing default warehouses: {str(e)}',
                'success': False
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ProductCategoryViewSet(viewsets.ModelViewSet):
    queryset = ProductCategory.objects.all()
    serializer_class = ProductCategorySerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['is_active']

class UnitCategoryViewSet(viewsets.ModelViewSet):
    queryset = UnitCategory.objects.all()
    serializer_class = UnitCategorySerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    authentication_classes = [TokenAuthentication]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['is_active']

class UnitCategoryListAPIView(ListAPIView):
    queryset = UnitCategory.objects.all()
    serializer_class = UnitCategorySerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [TokenAuthentication]

class ProductBrandViewSet(viewsets.ModelViewSet):
    queryset = ProductBrand.objects.all()
    serializer_class = ProductBrandSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['is_active', 'is_default']
    
class PriceTypeViewSet(viewsets.ModelViewSet):
    queryset = PriceType.objects.all()
    serializer_class = PriceTypeSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [TokenAuthentication]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['is_active']

class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['is_active', 'category', 'brands']
    search_fields = ['name', 'sku']
    ordering_fields = ['name', 'sku', 'created_at']
    ordering = ['name']

    def get_serializer_class(self):
        if self.action == 'list':
            return ProductListSerializer  # Para vistas tipo tabla
        elif self.action == 'retrieve':
            return ProductDetailSerializer  # Para vista detalle o edición
        return ProductSerializer  # Para create, update, partial_update

@permission_classes([AllowAny])
class ProductListAPIView(APIView):
    def get(self, request):
        # Filtrar solo productos activos por defecto
        is_active = request.query_params.get('is_active', 'true')
        if is_active.lower() == 'true':
            products = Product.objects.filter(is_active=True)
        elif is_active.lower() == 'false':
            products = Product.objects.filter(is_active=False)
        else:
            products = Product.objects.all()
            
        return Response([
            {"value": p.id, "label": p.name} for p in products
        ])

class UnitOfMeasureViewSet(viewsets.ModelViewSet):
    queryset = UnitOfMeasure.objects.all()
    serializer_class = UnitOfMeasureSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['is_active']

@permission_classes([AllowAny])
class UnitOfMeasureListAPIView(APIView):
    def get(self, request):
        units = UnitOfMeasure.objects.all()
        return Response([
            {"value": u.id, "label": u.name} for u in units
        ])

class ProductDataTableAPIView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        queryset = Product.objects.select_related('category', 'unit_default').prefetch_related('brands')
        return handle_datatable_query(
            request,
            queryset,
            ProductListSerializer,
            search_fields=['name', 'sku']
        )

class ProductListProviderAPIView(APIView):
    """
    Endpoint para provider pattern con server-side pagination, filtering y sorting
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Parámetros de paginación
            page = int(request.query_params.get('page', 1))
            per_page = int(request.query_params.get('per_page', 25))
            
            # Parámetros de filtrado
            is_active = request.query_params.get('is_active')
            search = request.query_params.get('search', '').strip()
            ordering = request.query_params.get('ordering', '-id')  # Descendente por ID
            
            # Construir queryset base
            queryset = Product.objects.select_related('category', 'unit_default').prefetch_related('brands')
            
            # Aplicar filtros
            if is_active is not None:
                if is_active.lower() == 'true':
                    queryset = queryset.filter(is_active=True)
                elif is_active.lower() == 'false':
                    queryset = queryset.filter(is_active=False)
            
            # Aplicar búsqueda
            if search:
                queryset = queryset.filter(
                    Q(name__icontains=search) |
                    Q(sku__icontains=search) |
                    Q(category__name__icontains=search)
                )
            
            # Aplicar ordenamiento
            queryset = queryset.order_by(ordering)
            
            # Calcular totales para estadísticas
            total_count = queryset.count()
            active_count = queryset.filter(is_active=True).count()
            inactive_count = queryset.filter(is_active=False).count()
            
            # Aplicar paginación
            start = (page - 1) * per_page
            end = start + per_page
            paginated_queryset = queryset[start:end]
            
            # Serializar datos
            serializer = ProductListSerializer(paginated_queryset, many=True)
            
            return Response({
                'items': serializer.data,
                'totalRows': total_count,
                'stats': {
                    'total': total_count,
                    'active': active_count,
                    'inactive': inactive_count
                }
            })
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class ProductListDirectAPIView(APIView):
    """
    Endpoint que devuelve productos como array directo (sin paginación)
    para compatibilidad con frontend que espera array directo
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Parámetros de filtrado
            is_active = request.query_params.get('is_active')
            search = request.query_params.get('search', '').strip()
            #ordering = request.query_params.get('ordering', 'id')
            ordering = request.query_params.get('ordering', '-id')  # Descendente por ID
            
            # Construir queryset
            queryset = Product.objects.select_related('category', 'unit_default').prefetch_related('brands')
            
            # Aplicar filtros
            if is_active is not None:
                if is_active.lower() == 'true':
                    queryset = queryset.filter(is_active=True)
                elif is_active.lower() == 'false':
                    queryset = queryset.filter(is_active=False)
            
            # Aplicar búsqueda
            if search:
                queryset = queryset.filter(
                    Q(name__icontains=search) |
                    Q(sku__icontains=search) |
                    Q(category__name__icontains=search)
                )
            
            # Aplicar ordenamiento
            queryset = queryset.order_by(ordering)
            
            
            # Serializar datos
            serializer = ProductListSerializer(queryset, many=True)
            
            # Calcular estadísticas
            total_count = queryset.count()
            active_count = queryset.filter(is_active=True).count()
            inactive_count = queryset.filter(is_active=False).count()
            
            return Response(serializer.data)
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)

class ProductBrandsListAPIView(APIView):
    """
    Endpoint para obtener las marcas disponibles para un producto específico
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, product_id):
        try:
            product = Product.objects.get(id=product_id)
            
            # Obtener todas las marcas del producto con información de default
            brands_data = []
            default_brand = product.get_default_brand()
            
            for brand in product.brands.filter(is_active=True):
                brands_data.append({
                    'id': brand.id,
                    'name': brand.name,
                    'is_default': brand == default_brand
                })
            
            return Response({
                'brands': brands_data,
                'default_brand': {
                    'id': default_brand.id if default_brand else None,
                    'name': default_brand.name if default_brand else None
                }
            })
                
        except Product.DoesNotExist:
            return Response({'error': 'Producto no encontrado'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)

class ProductBrandsUpdateAPIView(APIView):
    """
    Endpoint para actualizar las marcas de un producto
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, product_id):
        try:
            product = Product.objects.get(id=product_id)
            brands_data = request.data.get('brands', [])
            
            # Validar que se proporcione al menos una marca
            if not brands_data:
                return Response({
                    'error': 'El producto debe tener al menos una marca asignada.'
                }, status=400)
            
            # Actualizar marcas
            brands = ProductBrand.objects.filter(id__in=brands_data)
            if brands.count() != len(brands_data):
                return Response({
                    'error': 'Una o más marcas no existen.'
                }, status=400)
            
            product.brands.set(brands)
            product.ensure_default_brand()
            
            return Response({
                'message': 'Marcas actualizadas exitosamente',
                'success': True
            })
                
        except Product.DoesNotExist:
            return Response({'error': 'Producto no encontrado'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)

class ProductDefaultPriceAPIView(APIView):
    """
    Endpoint para obtener el precio predeterminado de un producto
    Soporta selección de precio según tipo de documento (compra/venta)
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, product_id):
        try:
            # Permitir filtrar por brand_id si se especifica
            brand_id = request.query_params.get('brand_id')
            # Nuevo: permitir filtrar por document_type_id
            document_type_id = request.query_params.get('document_type_id')
            
            product = Product.objects.get(id=product_id)
            
            # Obtener marca default del producto
            default_brand = product.get_default_brand()
            requested_brand = None
            
            # Si se especifica brand_id, encontrar la marca solicitada
            if brand_id:
                try:
                    requested_brand = product.brands.get(id=brand_id, is_active=True)
                except ProductBrand.DoesNotExist:
                    pass
            
            # Usar la marca solicitada o la default
            brand_to_use = requested_brand or default_brand
            
            # Determinar qué precio usar según el tipo de documento
            default_price = None
            
            if document_type_id:
                try:
                    doc_type = DocumentType.objects.get(id=document_type_id)
                    
                    if doc_type.is_purchase:
                        # Para documentos de compra: tomar el primer precio marcado como Purchase
                        default_price = product.prices.filter(
                            is_purchase=True, 
                            is_active=True
                        ).order_by('id').first()
                    elif doc_type.is_sales:
                        # Para documentos de venta: tomar el precio marcado como default y sale
                        default_price = product.prices.filter(
                            is_sale=True,
                            is_default=True, 
                            is_active=True
                        ).first()
                        
                        # Si no hay precio default de venta, tomar el primero de venta
                        if not default_price:
                            default_price = product.prices.filter(
                                is_sale=True,
                                is_active=True
                            ).order_by('id').first()
                except DocumentType.DoesNotExist:
                    pass
            
            # Fallback: si no se especificó document_type o no se encontró precio específico
            if not default_price:
                # Buscar el precio predeterminado (is_default=True)
                default_price = product.prices.filter(is_default=True, is_active=True).first()
            
            if not default_price:
                # Si no hay precio predeterminado, tomar el primero disponible
                default_price = product.prices.filter(is_active=True).first()
            
            if default_price:
                return Response({
                    'unit': default_price.unit.id,
                    'unit_name': default_price.unit.name,
                    'unit_price': default_price.price,
                    'price_type': default_price.price_type.id,
                    'price_type_name': default_price.price_type.name,
                    'brand': brand_to_use.id if brand_to_use else None,
                    'brand_name': brand_to_use.name if brand_to_use else None,
                    'default_brand': {
                        'id': default_brand.id if default_brand else None,
                        'name': default_brand.name if default_brand else None
                    }
                })
            else:
                return Response({
                    'unit': None,
                    'unit_name': None,
                    'unit_price': 0,
                    'price_type': None,
                    'price_type_name': None,
                    'brand': brand_to_use.id if brand_to_use else None,
                    'brand_name': brand_to_use.name if brand_to_use else None,
                    'default_brand': {
                        'id': default_brand.id if default_brand else None,
                        'name': default_brand.name if default_brand else None
                    }
                })
                
        except Product.DoesNotExist:
            return Response({'error': 'Product not found'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)

class DefaultWarehouseAPIView(APIView):
    """
    Endpoint para obtener el warehouse predeterminado
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            default_warehouse = Warehouse.objects.filter(is_default=True, is_active=True).first()
            
            if default_warehouse:
                return Response({
                    'id': default_warehouse.id,
                    'name': default_warehouse.name,
                })
            else:
                return Response({
                    'id': None,
                    'name': None,
                })
                
        except Exception as e:
            return Response({'error': str(e)}, status=500)


# ===== NUEVAS VISTAS PARA DASHBOARD MEJORADO =====

class TestDashboardAPIView(APIView):
    """
    Vista de prueba para verificar que las URLs funcionan
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response({'message': 'Dashboard API is working!'})


class InventoryDashboardMetricsAPIView(APIView):
    """
    Métricas generales del dashboard de inventario
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Métricas básicas
            total_products = Product.objects.filter(is_active=True).count()
            total_warehouses = Warehouse.objects.filter(is_active=True).count()
            total_stock_units = Stock.objects.aggregate(total=Sum('quantity'))['total'] or 0
            
            # Calcular valor total del inventario
            total_inventory_value = 0
            for stock in Stock.objects.select_related('product').all():
                default_price = stock.product.prices.filter(is_default=True, is_active=True).first()
                if default_price:
                    total_inventory_value += float(stock.quantity) * float(default_price.price)
            
            # Productos con stock bajo
            low_stock_count = Product.objects.filter(is_active=True).annotate(
                total_stock=Subquery(
                    Stock.objects.filter(product=OuterRef('pk'))
                    .values('product')
                    .annotate(qty=Sum('quantity'))
                    .values('qty')[:1]
                )
            ).filter(total_stock__lt=F('reorder_level')).count()
            
            # Productos vendidos en últimos 30 días
            from datetime import datetime, timedelta
            thirty_days_ago = datetime.now() - timedelta(days=30)
            products_sold = DocumentLine.objects.filter(
                document__document_type__is_sales=True,
                document__date__gte=thirty_days_ago.date(),
                document__is_active=True
            ).values('product').distinct().count()
            
            # Proveedores y clientes activos
            from apptransactions.models import Party
            active_suppliers = Party.objects.filter(supplier_rank__gt=0, is_active=True).count()
            active_customers = Party.objects.filter(customer_rank__gt=0, is_active=True).count()
            
            return Response({
                'total_products': total_products,
                'total_warehouses': total_warehouses,
                'total_stock_units': total_stock_units,
                'total_inventory_value': total_inventory_value,
                'low_stock_count': low_stock_count,
                'products_sold': products_sold,
                'active_suppliers': active_suppliers,
                'active_customers': active_customers
            })
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class TopSellingProductsAPIView(APIView):
    """
    Top 25 productos más vendidos por período
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            period_days = int(request.query_params.get('period_days', 30))
            from datetime import datetime, timedelta
            start_date = datetime.now() - timedelta(days=period_days)
            
            # Agregación de ventas por producto
            top_products = DocumentLine.objects.filter(
                document__document_type__is_sales=True,
                document__date__gte=start_date.date(),
                document__is_active=True
            ).values('product').annotate(
                quantity_sold=Sum('quantity'),
                total_value=Sum('final_price'),
                transaction_count=Count('document', distinct=True),
                last_sale_date=Max('document__date')
            ).order_by('-total_value')[:25]
            
            # Enriquecer con datos del producto
            result = []
            for item in top_products:
                product = Product.objects.select_related('category').get(id=item['product'])
                result.append({
                    'id': product.id,
                    'name': product.name,
                    'sku': product.sku,
                    'category': {
                        'id': product.category.id if product.category else None,
                        'name': product.category.name if product.category else None
                    },
                    'quantity_sold': item['quantity_sold'] or 0,
                    'total_value': float(item['total_value'] or 0),
                    'transaction_count': item['transaction_count'] or 0,
                    'last_sale_date': item['last_sale_date']
                })
            
            return Response(result)
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class SalesAnalysisAPIView(APIView):
    """
    Análisis completo de ventas con métricas y top productos
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            period_days = int(request.query_params.get('period_days', 30))
            from datetime import datetime, timedelta
            start_date = datetime.now() - timedelta(days=period_days)
            
            # Top productos más vendidos
            top_products = DocumentLine.objects.filter(
                document__document_type__is_sales=True,
                document__date__gte=start_date.date(),
                document__is_active=True
            ).values('product').annotate(
                quantity_sold=Sum('quantity'),
                total_value=Sum('final_price'),
                transaction_count=Count('document', distinct=True),
                last_sale_date=Max('document__date')
            ).order_by('-total_value')[:25]
            
            # Enriquecer con datos del producto
            top_products_list = []
            for item in top_products:
                product = Product.objects.select_related('category').get(id=item['product'])
                top_products_list.append({
                    'id': product.id,
                    'name': product.name,
                    'sku': product.sku,
                    'category': {
                        'id': product.category.id if product.category else None,
                        'name': product.category.name if product.category else None
                    },
                    'quantity_sold': item['quantity_sold'] or 0,
                    'total_value': float(item['total_value'] or 0),
                    'transaction_count': item['transaction_count'] or 0,
                    'last_sale_date': item['last_sale_date']
                })
            
            # Métricas de ventas
            total_sales = DocumentLine.objects.filter(
                document__document_type__is_sales=True,
                document__date__gte=start_date.date(),
                document__is_active=True
            ).aggregate(
                total_sales=Sum('final_price'),
                total_quantity=Sum('quantity'),
                transaction_count=Count('document', distinct=True)
            )
            
            # Cálculo de crecimiento (comparar con período anterior)
            previous_start = start_date - timedelta(days=period_days)
            previous_sales = DocumentLine.objects.filter(
                document__document_type__is_sales=True,
                document__date__gte=previous_start.date(),
                document__date__lt=start_date.date(),
                document__is_active=True
            ).aggregate(total_sales=Sum('final_price'))['total_sales'] or 0
            
            current_sales = total_sales['total_sales'] or 0
            growth_percentage = 0
            if previous_sales > 0:
                growth_percentage = ((current_sales - previous_sales) / previous_sales) * 100
            
            metrics = {
                'total_sales': float(current_sales),
                'total_quantity': total_sales['total_quantity'] or 0,
                'transaction_count': total_sales['transaction_count'] or 0,
                'daily_average': float(current_sales) / period_days if period_days > 0 else 0,
                'growth_percentage': round(growth_percentage, 2)
            }
            
            return Response({
                'topSellingProducts': top_products_list,
                'metrics': metrics
            })
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class LowStockProductsAPIView(APIView):
    """
    Productos con stock por debajo del nivel de reorden
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Productos con stock bajo (por debajo del reorder_level)
            low_stock_products = Product.objects.filter(
                is_active=True
            ).annotate(
                total_stock=Subquery(
                    Stock.objects.filter(product=OuterRef('pk'))
                    .values('product')
                    .annotate(qty=Sum('quantity'))
                    .values('qty')[:1]
                )
            ).filter(
                total_stock__lt=F('reorder_level')
            ).select_related('category').prefetch_related('brands').order_by('total_stock')
            
            result = []
            for product in low_stock_products:
                # Obtener precio por defecto
                default_price = product.prices.filter(is_default=True, is_active=True).first()
                
                result.append({
                    'id': product.id,
                    'name': product.name,
                    'sku': product.sku,
                    'category': {
                        'id': product.category.id if product.category else None,
                        'name': product.category.name if product.category else None
                    },
                    'brand': {
                        'id': product.get_default_brand().id if product.get_default_brand() else None,
                        'name': product.get_default_brand().name if product.get_default_brand() else None
                    },
                    'current_stock': product.total_stock or 0,
                    'reorder_level': product.reorder_level,
                    'stock_difference': (product.total_stock or 0) - product.reorder_level,
                    'default_price': float(default_price.price) if default_price else 0.0,
                    'stock_value': float(product.total_stock or 0) * float(default_price.price) if default_price else 0.0
                })
            
            return Response(result)
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class LowestStockProductsAPIView(APIView):
    """
    Top 25 productos con menor stock
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Top 25 productos con menor stock
            lowest_stock_products = Product.objects.filter(
                is_active=True
            ).annotate(
                total_stock=Subquery(
                    Stock.objects.filter(product=OuterRef('pk'))
                    .values('product')
                    .annotate(qty=Sum('quantity'))
                    .values('qty')[:1]
                )
            ).filter(
                total_stock__isnull=False
            ).select_related('category').prefetch_related('brands').order_by('total_stock')[:25]
            
            result = []
            for product in lowest_stock_products:
                # Obtener precio por defecto
                default_price = product.prices.filter(is_default=True, is_active=True).first()
                
                result.append({
                    'id': product.id,
                    'name': product.name,
                    'sku': product.sku,
                    'category': {
                        'id': product.category.id if product.category else None,
                        'name': product.category.name if product.category else None
                    },
                    'brand': {
                        'id': product.get_default_brand().id if product.get_default_brand() else None,
                        'name': product.get_default_brand().name if product.get_default_brand() else None
                    },
                    'current_stock': product.total_stock or 0,
                    'reorder_level': product.reorder_level,
                    'stock_difference': (product.total_stock or 0) - product.reorder_level,
                    'default_price': float(default_price.price) if default_price else 0.0,
                    'stock_value': float(product.total_stock or 0) * float(default_price.price) if default_price else 0.0,
                    'is_low_stock': (product.total_stock or 0) < product.reorder_level
                })
            
            return Response(result)
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class TopCustomersAPIView(APIView):
    """
    Top 10 clientes por volumen de compras
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            period_days = int(request.query_params.get('period_days', 90))
            from datetime import datetime, timedelta
            start_date = datetime.now() - timedelta(days=period_days)
            
            # Agregación de ventas por cliente (builder)
            top_customers = Document.objects.filter(
                document_type__is_sales=True,
                date__gte=start_date.date(),
                is_active=True,
                builder__isnull=False
            ).values('builder').annotate(
                total_purchases=Sum('total_amount'),
                transaction_count=Count('id'),
                last_purchase=Max('date')
            ).order_by('-total_purchases')[:10]
            
            # Enriquecer con datos del builder
            result = []
            for item in top_customers:
                from ctrctsapp.models import Builder
                builder = Builder.objects.select_related('party').get(id=item['builder'])
                result.append({
                    'id': builder.id,
                    'name': builder.name,
                    'party': {
                        'id': builder.party.id if builder.party else None,
                        'name': builder.party.name if builder.party else None,
                        'rfc': builder.party.rfc if builder.party else None
                    },
                    'total_purchases': float(item['total_purchases'] or 0),
                    'transaction_count': item['transaction_count'] or 0,
                    'last_purchase': item['last_purchase']
                })
            
            return Response({
                'topCustomers': result,
                'metrics': {
                    'total_customers': len(result),
                    'total_purchases': sum(item['total_purchases'] for item in result),
                    'total_transactions': sum(item['transaction_count'] for item in result)
                }
            })
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class TopSuppliersAPIView(APIView):
    """
    Top 10 proveedores por volumen de compras
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            period_days = int(request.query_params.get('period_days', 90))
            from datetime import datetime, timedelta
            start_date = datetime.now() - timedelta(days=period_days)
            
            # Agregación de compras por proveedor (builder)
            top_suppliers = Document.objects.filter(
                document_type__is_purchase=True,
                date__gte=start_date.date(),
                is_active=True,
                builder__isnull=False
            ).values('builder').annotate(
                total_purchases=Sum('total_amount'),
                transaction_count=Count('id'),
                last_purchase=Max('date')
            ).order_by('-total_purchases')[:10]
            
            # Enriquecer con datos del builder
            result = []
            for item in top_suppliers:
                from ctrctsapp.models import Builder
                builder = Builder.objects.select_related('party').get(id=item['builder'])
                result.append({
                    'id': builder.id,
                    'name': builder.name,
                    'party': {
                        'id': builder.party.id if builder.party else None,
                        'name': builder.party.name if builder.party else None,
                        'rfc': builder.party.rfc if builder.party else None
                    },
                    'total_purchases': float(item['total_purchases'] or 0),
                    'transaction_count': item['transaction_count'] or 0,
                    'last_purchase': item['last_purchase']
                })
            
            return Response({
                'topSuppliers': result,
                'metrics': {
                    'total_suppliers': len(result),
                    'total_purchases': sum(item['total_purchases'] for item in result),
                    'total_transactions': sum(item['transaction_count'] for item in result)
                }
            })
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class CustomersSuppliersComparisonAPIView(APIView):
    """
    Comparación de volumen entre clientes y proveedores por mes
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from datetime import datetime, timedelta
            from django.db.models.functions import TruncMonth
            
            # Últimos 12 meses
            twelve_months_ago = datetime.now() - timedelta(days=365)
            
            # Ventas por mes
            sales_by_month = Document.objects.filter(
                document_type__is_sales=True,
                date__gte=twelve_months_ago.date(),
                is_active=True
            ).annotate(
                month=TruncMonth('date')
            ).values('month').annotate(
                total_sales=Sum('total_amount')
            ).order_by('month')
            
            # Compras por mes
            purchases_by_month = Document.objects.filter(
                document_type__is_purchase=True,
                date__gte=twelve_months_ago.date(),
                is_active=True
            ).annotate(
                month=TruncMonth('date')
            ).values('month').annotate(
                total_purchases=Sum('total_amount')
            ).order_by('month')
            
            # Crear diccionario para facilitar la combinación
            sales_dict = {item['month'].strftime('%Y-%m'): float(item['total_sales'] or 0) for item in sales_by_month}
            purchases_dict = {item['month'].strftime('%Y-%m'): float(item['total_purchases'] or 0) for item in purchases_by_month}
            
            # Generar datos para los últimos 12 meses
            result = []
            for i in range(12):
                date = datetime.now() - timedelta(days=30*i)
                month_key = date.strftime('%Y-%m')
                month_name = date.strftime('%b')
                
                result.append({
                    'month': month_name,
                    'sales': sales_dict.get(month_key, 0),
                    'purchases': purchases_dict.get(month_key, 0)
                })
            
            result.reverse()  # Ordenar cronológicamente
            
            return Response(result)
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class ProductMovementsReportAPIView(APIView):
    """
    Reporte de movimientos de productos con filtros
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from datetime import datetime, timedelta
            from django.utils import timezone
            
            # Filtros
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            document_type = request.query_params.get('document_type')
            warehouse_id = request.query_params.get('warehouse_id')
            
            # Fechas por defecto (últimos 30 días)
            if not start_date:
                start_date = (timezone.now() - timedelta(days=30)).date()
            else:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                
            if not end_date:
                end_date = timezone.now().date()
            else:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            # Convertir fechas a datetime para incluir todo el día
            # Usar un enfoque más simple y directo
            start_datetime = datetime.combine(start_date, datetime.min.time())
            end_datetime = datetime.combine(end_date, datetime.max.time())
            
            # Query base con filtro de datetime preciso
            movements = InventoryMovement.objects.filter(
                timestamp__gte=start_datetime,
                timestamp__lte=end_datetime
            ).select_related('product', 'warehouse', 'created_by')
            
            # Aplicar filtros adicionales
            if document_type:
                # Buscar los IDs de documentos que correspondan al tipo de documento seleccionado
                from apptransactions.models import DocumentType, Document
                try:
                    doc_type_obj = DocumentType.objects.get(id=document_type)
                    
                    # Obtener todos los documentos de este tipo
                    document_ids = Document.objects.filter(
                        document_type_id=document_type,
                        is_active=True
                    ).values_list('id', flat=True)
                    
                    # Filtrar movimientos por IDs de documentos
                    movements = movements.filter(document__in=[str(doc_id) for doc_id in document_ids])
                    # Log estratégico único
                    print(f"ProductMovements: Filtered by {doc_type_obj.type_code}, found {movements.count()} movements")
                except DocumentType.DoesNotExist:
                    pass
            if warehouse_id:
                movements = movements.filter(warehouse_id=warehouse_id)
            
            # Ordenar por fecha más reciente
            movements = movements.order_by('-timestamp')[:100]  # Limitar a 100 registros
            
            # Serializar datos
            result = []
            for movement in movements:
                result.append({
                    'id': movement.id,
                    'date': movement.timestamp.date(),
                    'time': movement.timestamp.time(),
                    'product': {
                        'id': movement.product.id,
                        'name': movement.product.name,
                        'sku': movement.product.sku
                    },
                    'warehouse': {
                        'id': movement.warehouse.id,
                        'name': movement.warehouse.name
                    },
                    'quantity': float(movement.quantity),
                    'movement_type': movement.get_movement_type_display(),
                    'reason': movement.reason,
                    'document': movement.document,
                    'created_by': {
                        'id': movement.created_by.id if movement.created_by else None,
                        'username': movement.created_by.username if movement.created_by else 'Sistema'
                    }
                })
            
            return Response(result)
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class ProductMovementsExportAPIView(APIView):
    """
    Export product movements report to Excel
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from datetime import datetime, timedelta
            from django.utils import timezone
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
            
            # Get filters
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            document_type = request.query_params.get('document_type')
            warehouse_id = request.query_params.get('warehouse_id')
            
            # Default dates (last 30 days)
            if not start_date:
                start_date = (timezone.now() - timedelta(days=30)).date()
            else:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                
            if not end_date:
                end_date = timezone.now().date()
            else:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            # Convert dates to datetime to include the entire day
            # Use a simpler and more direct approach
            start_datetime = datetime.combine(start_date, datetime.min.time())
            end_datetime = datetime.combine(end_date, datetime.max.time())
            
            # Base query with precise datetime filter
            movements = InventoryMovement.objects.filter(
                timestamp__gte=start_datetime,
                timestamp__lte=end_datetime
            ).select_related('product', 'warehouse', 'created_by')
            
            # Apply additional filters
            if document_type:
                # Buscar los IDs de documentos que correspondan al tipo de documento seleccionado
                from apptransactions.models import DocumentType, Document
                try:
                    doc_type_obj = DocumentType.objects.get(id=document_type)
                    
                    # Obtener todos los documentos de este tipo
                    document_ids = Document.objects.filter(
                        document_type_id=document_type,
                        is_active=True
                    ).values_list('id', flat=True)
                    
                    # Filtrar movimientos por IDs de documentos
                    movements = movements.filter(document__in=[str(doc_id) for doc_id in document_ids])
                except DocumentType.DoesNotExist:
                    pass  # Si no existe el tipo de documento, no aplicar filtro
            if warehouse_id:
                movements = movements.filter(warehouse_id=warehouse_id)
            
            # Order by most recent
            movements = movements.order_by('-timestamp')
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Product Movements Report"
            
            # Headers
            headers = ['Date', 'Time', 'Type', 'Product Name', 'SKU', 'Quantity', 'Warehouse', 'Document', 'User', 'Reason']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row, movement in enumerate(movements, 2):
                ws.cell(row=row, column=1, value=movement.timestamp.date())
                ws.cell(row=row, column=2, value=movement.timestamp.time())
                ws.cell(row=row, column=3, value=movement.get_movement_type_display())
                ws.cell(row=row, column=4, value=movement.product.name)
                ws.cell(row=row, column=5, value=movement.product.sku)
                ws.cell(row=row, column=6, value=float(movement.quantity))
                ws.cell(row=row, column=7, value=movement.warehouse.name)
                ws.cell(row=row, column=8, value=movement.document or 'N/A')
                ws.cell(row=row, column=9, value=movement.created_by.username if movement.created_by else 'System')
                ws.cell(row=row, column=10, value=movement.reason or '')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="product_movements_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


# ===== EXPORT VIEWS =====

class SalesByProductExportAPIView(APIView):
    """
    Export sales data by product to Excel
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from datetime import datetime, timedelta
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
            
            period_days = int(request.query_params.get('period_days', 30))
            start_date = datetime.now() - timedelta(days=period_days)
            
            # Get sales data by product
            sales_data = DocumentLine.objects.filter(
                document__document_type__is_sales=True,
                document__date__gte=start_date.date(),
                document__is_active=True
            ).values('product').annotate(
                product_name=F('product__name'),
                product_sku=F('product__sku'),
                total_quantity=Sum('quantity'),
                total_value=Sum('final_price'),
                transaction_count=Count('document', distinct=True)
            ).order_by('-total_value')
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales by Product"
            
            # Headers
            headers = ['Product Name', 'SKU', 'Total Quantity', 'Total Value', 'Transactions']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row, item in enumerate(sales_data, 2):
                ws.cell(row=row, column=1, value=item['product_name'])
                ws.cell(row=row, column=2, value=item['product_sku'])
                ws.cell(row=row, column=3, value=float(item['total_quantity'] or 0))
                ws.cell(row=row, column=4, value=float(item['total_value'] or 0))
                ws.cell(row=row, column=5, value=item['transaction_count'] or 0)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="sales_by_product_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class SalesByCustomerExportAPIView(APIView):
    """
    Export sales data by customer to Excel
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from datetime import datetime, timedelta
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
            
            period_days = int(request.query_params.get('period_days', 30))
            start_date = datetime.now() - timedelta(days=period_days)
            
            # Get sales data by customer
            sales_data = Document.objects.filter(
                document_type__is_sales=True,
                date__gte=start_date.date(),
                is_active=True,
                builder__isnull=False
            ).values('builder').annotate(
                customer_name=F('builder__name'),
                total_purchases=Sum('total_amount'),
                transaction_count=Count('id'),
                last_purchase=Max('date')
            ).order_by('-total_purchases')
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales by Customer"
            
            # Headers
            headers = ['Customer Name', 'Total Purchases', 'Transactions', 'Last Purchase']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row, item in enumerate(sales_data, 2):
                ws.cell(row=row, column=1, value=item['customer_name'])
                ws.cell(row=row, column=2, value=float(item['total_purchases'] or 0))
                ws.cell(row=row, column=3, value=item['transaction_count'] or 0)
                ws.cell(row=row, column=4, value=item['last_purchase'].strftime('%Y-%m-%d') if item['last_purchase'] else '')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="sales_by_customer_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class SalesByPeriodExportAPIView(APIView):
    """
    Export sales data by period to Excel
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from datetime import datetime, timedelta
            from django.db.models.functions import TruncMonth
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
            
            # Get sales data by month for last 12 months
            twelve_months_ago = datetime.now() - timedelta(days=365)
            
            sales_data = Document.objects.filter(
                document_type__is_sales=True,
                date__gte=twelve_months_ago.date(),
                is_active=True
            ).annotate(
                month=TruncMonth('date')
            ).values('month').annotate(
                total_sales=Sum('total_amount'),
                transaction_count=Count('id')
            ).order_by('month')
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales by Period"
            
            # Headers
            headers = ['Month', 'Total Sales', 'Transactions']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row, item in enumerate(sales_data, 2):
                ws.cell(row=row, column=1, value=item['month'].strftime('%Y-%m'))
                ws.cell(row=row, column=2, value=float(item['total_sales'] or 0))
                ws.cell(row=row, column=3, value=item['transaction_count'] or 0)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="sales_by_period_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class FinancialSummaryExportAPIView(APIView):
    """
    Export financial summary to Excel
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from datetime import datetime, timedelta
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
            
            period_days = int(request.query_params.get('period_days', 30))
            start_date = datetime.now() - timedelta(days=period_days)
            
            # Get financial summary data
            sales_summary = DocumentLine.objects.filter(
                document__document_type__is_sales=True,
                document__date__gte=start_date.date(),
                document__is_active=True
            ).aggregate(
                total_sales=Sum('final_price'),
                total_quantity=Sum('quantity'),
                transaction_count=Count('document', distinct=True)
            )
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Financial Summary"
            
            # Headers
            headers = ['Metric', 'Value']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            data = [
                ['Total Sales', float(sales_summary['total_sales'] or 0)],
                ['Total Quantity Sold', sales_summary['total_quantity'] or 0],
                ['Total Transactions', sales_summary['transaction_count'] or 0],
                ['Average Sale Value', float(sales_summary['total_sales'] or 0) / max(sales_summary['transaction_count'] or 1, 1)],
                ['Period (Days)', period_days],
                ['Report Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            ]
            
            for row, (metric, value) in enumerate(data, 2):
                ws.cell(row=row, column=1, value=metric)
                ws.cell(row=row, column=2, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="financial_summary_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


# ===== STOCK EXPORT VIEWS =====

class StockByWarehouseExportAPIView(APIView):
    """
    Export stock data by warehouse to Excel
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from datetime import datetime
            import openpyxl
            from django.http import HttpResponse
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Stock by Warehouse"
            
            # Headers
            headers = ['Warehouse', 'Product', 'SKU', 'Category', 'Brand', 'Current Stock', 'Reorder Level', 'Status', 'Stock Value']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Obtener datos de stock por almacén
            from .models import Stock, Warehouse
            stock_data = Stock.objects.select_related('product', 'warehouse', 'product__category').prefetch_related('product__brands').all()
            
            row = 2
            for stock in stock_data:
                product = stock.product
                warehouse = stock.warehouse
                
                # Obtener precio por defecto
                default_price = product.prices.filter(is_default=True, is_active=True).first()
                stock_value = float(stock.quantity) * float(default_price.price) if default_price else 0.0
                
                # Determinar status
                status = "Low Stock" if stock.quantity < product.reorder_level else "OK"
                
                data = [
                    warehouse.name,
                    product.name,
                    product.sku,
                    product.category.name if product.category else 'N/A',
                    product.get_default_brand().name if product.get_default_brand() else 'N/A',
                    stock.quantity,
                    product.reorder_level,
                    status,
                    stock_value
                ]
                
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
                
                row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="stock_by_warehouse_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class CompleteStockExportAPIView(APIView):
    """
    Export complete stock inventory to Excel
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from datetime import datetime
            import openpyxl
            from django.http import HttpResponse
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Complete Stock"
            
            # Headers
            headers = ['Product', 'SKU', 'Category', 'Brand', 'Total Stock', 'Reorder Level', 'Status', 'Stock Value']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Obtener productos con stock total
            products = Product.objects.filter(is_active=True).annotate(
                total_stock=Subquery(
                    Stock.objects.filter(product=OuterRef('pk'))
                    .values('product')
                    .annotate(qty=Sum('quantity'))
                    .values('qty')[:1]
                )
            ).select_related('category').prefetch_related('brands').order_by('name')
            
            row = 2
            for product in products:
                # Obtener precio por defecto
                default_price = product.prices.filter(is_default=True, is_active=True).first()
                stock_value = float(product.total_stock or 0) * float(default_price.price) if default_price else 0.0
                
                # Determinar status
                status = "Low Stock" if (product.total_stock or 0) < product.reorder_level else "OK"
                
                data = [
                    product.name,
                    product.sku,
                    product.category.name if product.category else 'N/A',
                    product.get_default_brand().name if product.get_default_brand() else 'N/A',
                    product.total_stock or 0,
                    product.reorder_level,
                    status,
                    stock_value
                ]
                
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
                
                row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="complete_stock_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class LowStockProductsExportAPIView(APIView):
    """
    Export low stock products to Excel
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from datetime import datetime
            import openpyxl
            from django.http import HttpResponse
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Low Stock Products"
            
            # Headers
            headers = ['Product', 'SKU', 'Category', 'Brand', 'Current Stock', 'Reorder Level', 'Stock Difference', 'Status', 'Stock Value']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Obtener productos con stock bajo
            low_stock_products = Product.objects.filter(
                is_active=True
            ).annotate(
                total_stock=Subquery(
                    Stock.objects.filter(product=OuterRef('pk'))
                    .values('product')
                    .annotate(qty=Sum('quantity'))
                    .values('qty')[:1]
                )
            ).filter(
                total_stock__lt=F('reorder_level')
            ).select_related('category').prefetch_related('brands').order_by('total_stock')
            
            row = 2
            for product in low_stock_products:
                # Obtener precio por defecto
                default_price = product.prices.filter(is_default=True, is_active=True).first()
                stock_value = float(product.total_stock or 0) * float(default_price.price) if default_price else 0.0
                
                # Determinar status
                status = "Out of Stock" if (product.total_stock or 0) == 0 else "Low Stock"
                stock_difference = (product.total_stock or 0) - product.reorder_level
                
                data = [
                    product.name,
                    product.sku,
                    product.category.name if product.category else 'N/A',
                    product.get_default_brand().name if product.get_default_brand() else 'N/A',
                    product.total_stock or 0,
                    product.reorder_level,
                    stock_difference,
                    status,
                    stock_value
                ]
                
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
                
                row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="low_stock_products_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class StockReportExportAPIView(APIView):
    """
    Export comprehensive stock report to Excel
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from datetime import datetime
            import openpyxl
            from django.http import HttpResponse
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Crear workbook
            wb = openpyxl.Workbook()
            
            # Hoja 1: Resumen de Stock
            ws1 = wb.active
            ws1.title = "Stock Summary"
            
            # Headers
            headers = ['Product', 'SKU', 'Category', 'Brand', 'Total Stock', 'Reorder Level', 'Status', 'Stock Value']
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Obtener todos los productos
            products = Product.objects.filter(is_active=True).annotate(
                total_stock=Subquery(
                    Stock.objects.filter(product=OuterRef('pk'))
                    .values('product')
                    .annotate(qty=Sum('quantity'))
                    .values('qty')[:1]
                )
            ).select_related('category').prefetch_related('brands').order_by('name')
            
            row = 2
            for product in products:
                # Obtener precio por defecto
                default_price = product.prices.filter(is_default=True, is_active=True).first()
                stock_value = float(product.total_stock or 0) * float(default_price.price) if default_price else 0.0
                
                # Determinar status
                status = "Low Stock" if (product.total_stock or 0) < product.reorder_level else "OK"
                
                data = [
                    product.name,
                    product.sku,
                    product.category.name if product.category else 'N/A',
                    product.get_default_brand().name if product.get_default_brand() else 'N/A',
                    product.total_stock or 0,
                    product.reorder_level,
                    status,
                    stock_value
                ]
                
                for col, value in enumerate(data, 1):
                    ws1.cell(row=row, column=col, value=value)
                
                row += 1
            
            # Hoja 2: Stock por Almacén
            ws2 = wb.create_sheet("Stock by Warehouse")
            
            # Headers
            headers2 = ['Warehouse', 'Product', 'SKU', 'Current Stock', 'Reorder Level', 'Status']
            for col, header in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Obtener datos de stock por almacén
            stock_data = Stock.objects.select_related('product', 'warehouse').all()
            
            row = 2
            for stock in stock_data:
                product = stock.product
                warehouse = stock.warehouse
                
                # Determinar status
                status = "Low Stock" if stock.quantity < product.reorder_level else "OK"
                
                data = [
                    warehouse.name,
                    product.name,
                    product.sku,
                    stock.quantity,
                    product.reorder_level,
                    status
                ]
                
                for col, value in enumerate(data, 1):
                    ws2.cell(row=row, column=col, value=value)
                
                row += 1
            
            # Auto-adjust column widths para ambas hojas
            for ws in [ws1, ws2]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="stock_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


# ===== CUSTOMERS & SUPPLIERS EXPORT VIEWS =====

class CustomersListExportAPIView(APIView):
    """
    Export customers list to Excel
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from datetime import datetime
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
            from ctrctsapp.models import Builder
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Customers List"
            
            # Headers
            headers = ['Customer Name', 'Party Name', 'RFC', 'Total Purchases', 'Transactions', 'Last Purchase']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Get customers data (builders with sales)
            customers_data = Document.objects.filter(
                document_type__is_sales=True,
                is_active=True,
                builder__isnull=False
            ).values('builder').annotate(
                customer_name=F('builder__name'),
                party_name=F('builder__party__name'),
                rfc=F('builder__party__rfc'),
                total_purchases=Sum('total_amount'),
                transaction_count=Count('id'),
                last_purchase=Max('date')
            ).order_by('-total_purchases')
            
            # Data
            for row, item in enumerate(customers_data, 2):
                ws.cell(row=row, column=1, value=item['customer_name'])
                ws.cell(row=row, column=2, value=item['party_name'])
                ws.cell(row=row, column=3, value=item['rfc'])
                ws.cell(row=row, column=4, value=float(item['total_purchases'] or 0))
                ws.cell(row=row, column=5, value=item['transaction_count'] or 0)
                ws.cell(row=row, column=6, value=item['last_purchase'].strftime('%Y-%m-%d') if item['last_purchase'] else '')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="customers_list_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class SuppliersListExportAPIView(APIView):
    """
    Export suppliers list to Excel
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from datetime import datetime
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
            from ctrctsapp.models import Builder
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Suppliers List"
            
            # Headers
            headers = ['Supplier Name', 'Party Name', 'RFC', 'Total Purchases', 'Transactions', 'Last Purchase']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Get suppliers data (builders with purchases)
            suppliers_data = Document.objects.filter(
                document_type__is_purchase=True,
                is_active=True,
                builder__isnull=False
            ).values('builder').annotate(
                supplier_name=F('builder__name'),
                party_name=F('builder__party__name'),
                rfc=F('builder__party__rfc'),
                total_purchases=Sum('total_amount'),
                transaction_count=Count('id'),
                last_purchase=Max('date')
            ).order_by('-total_purchases')
            
            # Data
            for row, item in enumerate(suppliers_data, 2):
                ws.cell(row=row, column=1, value=item['supplier_name'])
                ws.cell(row=row, column=2, value=item['party_name'])
                ws.cell(row=row, column=3, value=item['rfc'])
                ws.cell(row=row, column=4, value=float(item['total_purchases'] or 0))
                ws.cell(row=row, column=5, value=item['transaction_count'] or 0)
                ws.cell(row=row, column=6, value=item['last_purchase'].strftime('%Y-%m-%d') if item['last_purchase'] else '')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="suppliers_list_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class ComparativeAnalysisExportAPIView(APIView):
    """
    Export comparative analysis between customers and suppliers to Excel
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from datetime import datetime
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            
            # Sheet 1: Sales vs Purchases Comparison
            ws1 = wb.active
            ws1.title = "Sales vs Purchases"
            
            # Headers
            headers = ['Month', 'Total Sales', 'Total Purchases', 'Net Profit', 'Growth Rate']
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Get monthly data for last 12 months
            from datetime import timedelta
            from django.db.models.functions import TruncMonth
            
            twelve_months_ago = datetime.now() - timedelta(days=365)
            
            # Sales by month
            sales_by_month = Document.objects.filter(
                document_type__is_sales=True,
                date__gte=twelve_months_ago.date(),
                is_active=True
            ).annotate(
                month=TruncMonth('date')
            ).values('month').annotate(
                total_sales=Sum('total_amount')
            ).order_by('month')
            
            # Purchases by month
            purchases_by_month = Document.objects.filter(
                document_type__is_purchase=True,
                date__gte=twelve_months_ago.date(),
                is_active=True
            ).annotate(
                month=TruncMonth('date')
            ).values('month').annotate(
                total_purchases=Sum('total_amount')
            ).order_by('month')
            
            # Create dictionaries for easy lookup
            sales_dict = {item['month'].strftime('%Y-%m'): float(item['total_sales'] or 0) for item in sales_by_month}
            purchases_dict = {item['month'].strftime('%Y-%m'): float(item['total_purchases'] or 0) for item in purchases_by_month}
            
            # Generate data for last 12 months
            row = 2
            previous_sales = 0
            for i in range(12):
                date = datetime.now() - timedelta(days=30*i)
                month_key = date.strftime('%Y-%m')
                month_name = date.strftime('%b %Y')
                
                sales = sales_dict.get(month_key, 0)
                purchases = purchases_dict.get(month_key, 0)
                net_profit = sales - purchases
                
                # Calculate growth rate
                growth_rate = 0
                if previous_sales > 0:
                    growth_rate = ((sales - previous_sales) / previous_sales) * 100
                
                data = [month_name, sales, purchases, net_profit, f"{growth_rate:.1f}%"]
                for col, value in enumerate(data, 1):
                    ws1.cell(row=row, column=col, value=value)
                
                previous_sales = sales
                row += 1
            
            # Sheet 2: Top Customers vs Top Suppliers
            ws2 = wb.create_sheet("Top Customers vs Suppliers")
            
            # Headers
            headers2 = ['Rank', 'Customer/Supplier', 'Type', 'Total Amount', 'Transactions', 'Last Activity']
            for col, header in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Get top customers
            top_customers = Document.objects.filter(
                document_type__is_sales=True,
                date__gte=twelve_months_ago.date(),
                is_active=True,
                builder__isnull=False
            ).values('builder').annotate(
                name=F('builder__name'),
                total_amount=Sum('total_amount'),
                transaction_count=Count('id'),
                last_activity=Max('date')
            ).order_by('-total_amount')[:10]
            
            # Get top suppliers
            top_suppliers = Document.objects.filter(
                document_type__is_purchase=True,
                date__gte=twelve_months_ago.date(),
                is_active=True,
                builder__isnull=False
            ).values('builder').annotate(
                name=F('builder__name'),
                total_amount=Sum('total_amount'),
                transaction_count=Count('id'),
                last_activity=Max('date')
            ).order_by('-total_amount')[:10]
            
            row = 2
            # Add customers
            for idx, customer in enumerate(top_customers, 1):
                data = [
                    idx,
                    customer['name'],
                    'Customer',
                    float(customer['total_amount'] or 0),
                    customer['transaction_count'] or 0,
                    customer['last_activity'].strftime('%Y-%m-%d') if customer['last_activity'] else ''
                ]
                for col, value in enumerate(data, 1):
                    ws2.cell(row=row, column=col, value=value)
                row += 1
            
            # Add suppliers
            for idx, supplier in enumerate(top_suppliers, 1):
                data = [
                    idx,
                    supplier['name'],
                    'Supplier',
                    float(supplier['total_amount'] or 0),
                    supplier['transaction_count'] or 0,
                    supplier['last_activity'].strftime('%Y-%m-%d') if supplier['last_activity'] else ''
                ]
                for col, value in enumerate(data, 1):
                    ws2.cell(row=row, column=col, value=value)
                row += 1
            
            # Auto-adjust column widths for both sheets
            for ws in [ws1, ws2]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="comparative_analysis_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class TrendsExportAPIView(APIView):
    """
    Export monthly trends analysis to Excel
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from datetime import datetime
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Monthly Trends"
            
            # Headers
            headers = ['Month', 'Sales', 'Purchases', 'Net Profit', 'Growth %', 'Transactions', 'Avg Transaction']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Get monthly trends data
            from datetime import timedelta
            from django.db.models.functions import TruncMonth
            
            twelve_months_ago = datetime.now() - timedelta(days=365)
            
            # Sales by month
            sales_by_month = Document.objects.filter(
                document_type__is_sales=True,
                date__gte=twelve_months_ago.date(),
                is_active=True
            ).annotate(
                month=TruncMonth('date')
            ).values('month').annotate(
                total_sales=Sum('total_amount'),
                transaction_count=Count('id')
            ).order_by('month')
            
            # Purchases by month
            purchases_by_month = Document.objects.filter(
                document_type__is_purchase=True,
                date__gte=twelve_months_ago.date(),
                is_active=True
            ).annotate(
                month=TruncMonth('date')
            ).values('month').annotate(
                total_purchases=Sum('total_amount')
            ).order_by('month')
            
            # Create dictionaries for easy lookup
            sales_dict = {item['month'].strftime('%Y-%m'): {
                'sales': float(item['total_sales'] or 0),
                'transactions': item['transaction_count'] or 0
            } for item in sales_by_month}
            
            purchases_dict = {item['month'].strftime('%Y-%m'): float(item['total_purchases'] or 0) for item in purchases_by_month}
            
            # Generate data for last 12 months
            row = 2
            previous_sales = 0
            for i in range(12):
                date = datetime.now() - timedelta(days=30*i)
                month_key = date.strftime('%Y-%m')
                month_name = date.strftime('%b %Y')
                
                sales_data = sales_dict.get(month_key, {'sales': 0, 'transactions': 0})
                sales = sales_data['sales']
                transactions = sales_data['transactions']
                purchases = purchases_dict.get(month_key, 0)
                net_profit = sales - purchases
                
                # Calculate growth rate
                growth_rate = 0
                if previous_sales > 0:
                    growth_rate = ((sales - previous_sales) / previous_sales) * 100
                
                # Calculate average transaction
                avg_transaction = sales / max(transactions, 1) if transactions > 0 else 0
                
                data = [
                    month_name,
                    sales,
                    purchases,
                    net_profit,
                    f"{growth_rate:.1f}%",
                    transactions,
                    avg_transaction
                ]
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
                
                previous_sales = sales
                row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="trends_analysis_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)